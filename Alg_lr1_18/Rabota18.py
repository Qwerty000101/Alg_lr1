import openpyxl
wb = openpyxl.load_workbook('zadanie18_4.xlsx')
sheet = wb.active
for i in range(1, 11):
    for j in range(1, 11):
        sheet.cell(row=i+12, column=j).value = sheet.cell(row=i, column=j).value
for i in range(14,23):
    sheet[i][9].value=int(sheet[i-12][9].value)+int(sheet[i-1][9].value)
for i in range(8,0-1,-1):
    sheet[13][i].value=int(sheet[1][i].value)+int(sheet[13][i+1].value)
for i in range(14,23):
    for j in range(8,0-1,-1):
        sheet[i][j].value=int(sheet[i-12][j].value)+max(sheet[i-1][j].value,sheet[i][j+1].value)
rez1=sheet["A22"].value
for i in range(1, 11):
    for j in range(1, 11):
        sheet.cell(row=i+12, column=j).value = sheet.cell(row=i, column=j).value
for i in range(14,23):
    sheet[i][9].value=int(sheet[i-12][9].value)+int(sheet[i-1][9].value)
for i in range(8,0-1,-1):
    sheet[13][i].value=int(sheet[1][i].value)+int(sheet[13][i+1].value)
for i in range(14,23):
    for j in range(8,0-1,-1):
        sheet[i][j].value=int(sheet[i-12][j].value)+min(sheet[i-1][j].value,sheet[i][j+1].value)
rez2=sheet["A22"].value
print(rez1,rez2)
wb.save('zadanie18_4.xlsx')